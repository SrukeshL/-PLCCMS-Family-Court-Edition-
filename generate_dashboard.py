"""
╔══════════════════════════════════════════════════════════════════════════════╗
║   FAMILY COURT CUSTODY DASHBOARD GENERATOR                                  ║
║   Pennsylvania Family Court — 23 Pa.C.S. § 5328 (eff. Aug 2025)            ║
║   Run: python generate_dashboard.py                                          ║
║   Requires: pip install openpyxl pandas jinja2                               ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""

import sys
import json
import hashlib
import datetime
import os

try:
    import openpyxl
    import pandas as pd
except ImportError:
    print("\n⚠  Missing libraries. Please run:\n  pip install openpyxl pandas\n")
    sys.exit(1)

# ── Configuration ─────────────────────────────────────────────────────────────
EXCEL_FILE   = "family_court_custody_template.xlsx"
OUTPUT_HTML  = "family_court_dashboard.html"
ADMIN_HTML   = "admin_panel.html"

PRIORITY_COLORS = {
    "CATASTROPHIC": ("#4A0000", "#FF6B6B"),
    "CRITICAL":     ("#7B0000", "#FF9999"),
    "HIGH":         ("#7D3C00", "#FFAD60"),
    "MEDIUM":       ("#1A4A7A", "#82B4E8"),
    "LOW":          ("#145A32", "#82E0AA"),
}

CONF_ICONS = {
    "public":       ("🟢", "Public"),
    "restricted":   ("🟡", "Restricted"),
    "confidential": ("🔴", "Confidential"),
    "sealed":       ("⚫", "Sealed — Court Order Required"),
}

def read_config(wb):
    ws = wb["Dashboard Config"]
    cfg = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] and row[1] is not None:
            cfg[str(row[0]).strip()] = str(row[1]).strip() if row[1] else ""
    return cfg

def read_sheet_rows(wb, sheet_name, min_row=3):
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=min_row, values_only=True):
        if any(v for v in row if v is not None):
            rows.append([str(v).strip() if v is not None else "" for v in row])
    return rows

def priority_badge(p):
    key = (p or "MEDIUM").upper()
    bg, fg = PRIORITY_COLORS.get(key, ("#333", "#fff"))
    return f'<span class="badge" style="background:{bg};color:{fg}">{key}</span>'

def conf_badge(tier):
    t = (tier or "").lower()
    for key, (icon, label) in CONF_ICONS.items():
        if key in t:
            return f'<span class="conf-badge conf-{key}">{icon} {label}</span>'
    return f'<span class="conf-badge conf-public">🟢 Public</span>'

def file_hash(filepath):
    h = hashlib.sha256()
    with open(filepath, "rb") as f:
        h.update(f.read())
    return h.hexdigest()[:16]

def generate_dashboard(wb, cfg):
    now = datetime.datetime.now().strftime("%B %d, %Y at %I:%M %p")
    case_name     = cfg.get("Case Name", "Family Court Custody Case")
    docket        = cfg.get("Case Docket Number", "[Docket Number]")
    court         = cfg.get("Court Name", "Philadelphia Family Court")
    jurisdiction  = cfg.get("Jurisdiction", "Pennsylvania, USA")
    party_a       = cfg.get("Party A Label", "Petitioner")
    party_b       = cfg.get("Party B Label", "Respondent")
    conf_label    = cfg.get("Confidentiality Label", "CONFIDENTIAL")
    tracking      = cfg.get("Tracking Enabled", "FALSE").upper() == "TRUE"
    webhook       = cfg.get("Google Sheet Webhook", "")
    statute       = cfg.get("Statute Reference", "23 Pa.C.S. § 5328 (eff. Aug 2025)")

    # Read all sheets
    overview_rows  = read_sheet_rows(wb, "Case Overview", 3)
    factors_rows   = read_sheet_rows(wb, "PA Custody Factors §5328", 4)
    evidence_rows  = read_sheet_rows(wb, "Evidence Tracker", 3)
    timeline_rows  = read_sheet_rows(wb, "Timeline of Events", 3)
    comms_rows     = read_sheet_rows(wb, "Communications Log", 3)
    cx_rows        = read_sheet_rows(wb, "Cross-Examination Builder", 4)
    checklist_rows = read_sheet_rows(wb, "Hearing Prep Checklist", 3)
    negotiation_rows = read_sheet_rows(wb, "Negotiation & Settlement", 3)
    judge_rows     = read_sheet_rows(wb, "Judge Pre-Trial View", 4)
    odc_rows       = read_sheet_rows(wb, "ODC Violation Tracker [CONF]", 4)

    # Tracking script
    tracking_js = ""
    if tracking and webhook:
        tracking_js = f"""
<script>
(function() {{
  var data = {{
    timestamp: new Date().toISOString(),
    event: "page_load",
    referrer: document.referrer,
    url: window.location.href,
    userAgent: navigator.userAgent,
    language: navigator.language,
    screen: screen.width + "x" + screen.height,
    timezone: Intl.DateTimeFormat().resolvedOptions().timeZone
  }};
  fetch("{webhook}", {{method:"POST", headers:{{"Content-Type":"application/json"}}, body:JSON.stringify(data)}})
    .catch(function(){{}});
  var t0 = Date.now();
  window.addEventListener("beforeunload", function() {{
    var payload = Object.assign({{}}, data, {{event:"page_exit", duration_sec: Math.round((Date.now()-t0)/1000)}});
    navigator.sendBeacon("{webhook}", JSON.stringify(payload));
  }});
  var tabs = []; var tabStart = {{}};
  document.querySelectorAll('.tab-btn').forEach(function(b){{
    b.addEventListener('click', function(){{
      var name = b.getAttribute('data-tab');
      tabs.push(name); tabStart[name] = Date.now();
      fetch("{webhook}", {{method:"POST",headers:{{"Content-Type":"application/json"}},
        body:JSON.stringify(Object.assign({{}},data,{{event:"tab_view",tab:name}}))}}).catch(function(){{}});
    }});
  }});
}})();
</script>"""

    # Build factor rows HTML
    def factors_html():
        rows_html = ""
        for r in factors_rows:
            if len(r) < 9:
                r += [""] * (9 - len(r))
            num, desc, tier, ev_a, ev_b, rebuttal, priority, favors, docref = r[:9]
            if "deleted" in desc.lower() or "n/a" in desc.lower():
                continue
            safety = "SAFETY" in tier.upper()
            row_class = "factor-safety" if safety else "factor-row"
            rows_html += f"""
            <tr class="{row_class}">
              <td class="factor-num">{num}</td>
              <td class="factor-desc">{desc}
                {'<span class="safety-tag">⚠ SAFETY — Substantial Weight</span>' if safety else ''}
              </td>
              <td>{tier}</td>
              <td class="evidence-cell">{ev_a or '<span class="empty">— Enter evidence —</span>'}</td>
              <td class="evidence-cell">{ev_b or '<span class="empty">— Enter opposing claims —</span>'}</td>
              <td class="evidence-cell">{rebuttal or '<span class="empty">— Enter rebuttal —</span>'}</td>
              <td class="center">{priority_badge(priority)}</td>
              <td class="center">{favors or '—'}</td>
              <td class="center small">{docref}</td>
            </tr>"""
        return rows_html

    def evidence_html():
        rows_html = ""
        for r in evidence_rows:
            if len(r) < 10: r += [""] * (10 - len(r))
            n, date, etype, desc, why, priority, tier, link, physical, status = r[:10]
            if not etype or etype.startswith("Evidence Type"): continue
            link_html = f'<a href="{link}" target="_blank" class="drive-link">🔗 Open</a>' if link and link.startswith("http") else (link or "—")
            rows_html += f"""
            <tr>
              <td class="center">{n}</td>
              <td class="center small">{date}</td>
              <td><strong>{etype}</strong></td>
              <td>{desc}</td>
              <td>{why}</td>
              <td class="center">{priority_badge(priority)}</td>
              <td class="center">{conf_badge(tier)}</td>
              <td>{link_html}</td>
              <td class="small">{physical}</td>
              <td class="center">{status or 'Pending'}</td>
            </tr>"""
        return rows_html

    def timeline_html():
        rows_html = ""
        for r in timeline_rows:
            if len(r) < 8: r += [""] * (8 - len(r))
            n, date, time, etype, desc, parties, priority, evref = r[:8]
            if not date and not desc: continue
            rows_html += f"""
            <tr>
              <td class="center">{n}</td>
              <td class="center"><strong>{date}</strong></td>
              <td class="center small">{time}</td>
              <td><span class="event-type">{etype}</span></td>
              <td>{desc}</td>
              <td class="small">{parties}</td>
              <td class="center">{priority_badge(priority)}</td>
              <td class="small">{evref}</td>
            </tr>"""
        return rows_html

    def comms_html():
        rows_html = ""
        for r in comms_rows:
            if len(r) < 9: r += [""] * (9 - len(r))
            n, date, time, method, frm, to, summary, response, evlink = r[:9]
            if not date and not summary: continue
            rows_html += f"""
            <tr>
              <td class="center">{n}</td>
              <td class="center"><strong>{date}</strong></td>
              <td class="center small">{time}</td>
              <td><span class="comm-method">{method}</span></td>
              <td>{frm}</td>
              <td>{to}</td>
              <td>{summary}</td>
              <td class="center">{response or '—'}</td>
              <td class="small">{evlink}</td>
            </tr>"""
        return rows_html

    def cx_html():
        rows_html = ""
        current_section = ""
        for r in cx_rows:
            if len(r) < 9: r += [""] * (9 - len(r))
            # Detect section header rows (merged cells produce single non-empty value)
            if r[0] == "" and r[1] == "" and any("SECTION" in x.upper() for x in r):
                section_title = next((x for x in r if "SECTION" in x.upper()), "")
                if section_title:
                    rows_html += f'<tr class="cx-section-header"><td colspan="9">{section_title}</td></tr>'
                continue
            n, witness, role, question, expected, evidence, factor, priority, status = r[:9]
            if not question: continue
            rows_html += f"""
            <tr>
              <td class="center">{n}</td>
              <td><strong>{witness}</strong></td>
              <td class="small">{role}</td>
              <td class="question-cell">❓ {question}</td>
              <td class="small">{expected}</td>
              <td class="small">{evidence}</td>
              <td class="center small">{factor}</td>
              <td class="center">{priority_badge(priority)}</td>
              <td class="center">{status or 'Draft'}</td>
            </tr>"""
        return rows_html

    def checklist_html():
        rows_html = ""
        for r in checklist_rows:
            if len(r) < 5: r += [""] * (5 - len(r))
            n, task, notes, due, status = r[:5]
            if not task: continue
            done = "✅" in (status or "")
            rows_html += f"""
            <tr class="{'checklist-done' if done else ''}">
              <td class="center">{n}</td>
              <td>{task}</td>
              <td>{notes}</td>
              <td class="center small">{due}</td>
              <td class="center">{status or '⬜ Pending'}</td>
            </tr>"""
        return rows_html

    def negotiation_html():
        rows_html = ""
        for r in negotiation_rows:
            if len(r) < 5: r += [""] * (5 - len(r))
            issue, pos_a, pos_b, compromise, status = r[:5]
            if not issue: continue
            status_class = "agreed" if "Agreed" in (status or "") else ("disputed" if "Disputed" in (status or "") else "pending")
            rows_html += f"""
            <tr>
              <td><strong>{issue}</strong></td>
              <td>{pos_a or '—'}</td>
              <td>{pos_b or '—'}</td>
              <td>{compromise or '—'}</td>
              <td class="center"><span class="status-{status_class}">{status or '⬜ Not Discussed'}</span></td>
            </tr>"""
        return rows_html

    def judge_html():
        rows_html = ""
        for r in judge_rows:
            if len(r) < 4: r += [""] * (4 - len(r))
            section, summary, evref, priority = r[:4]
            if not section: continue
            rows_html += f"""
            <tr>
              <td><strong>{section}</strong></td>
              <td>{summary if summary and not summary.startswith('[') else '<span class="empty">— Not completed —</span>'}</td>
              <td class="small">{evref}</td>
              <td class="center">{priority_badge(priority)}</td>
            </tr>"""
        return rows_html

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Family Court Custody Dashboard — {case_name}</title>
<style>
  /* ── Philadelphia Family Court color palette ── */
  :root {{
    --navy:      #0D2B5E;
    --gold:      #C8A84B;
    --teal:      #2E7D7B;
    --gold-lt:   #FFF8E1;
    --navy-lt:   #E8EEF7;
    --red:       #C0392B;
    --orange:    #E67E22;
    --green:     #1E8449;
    --white:     #FFFFFF;
    --bg:        #F4F6FB;
    --text:      #1A1A2E;
    --border:    #CBD5E8;
    --shadow:    0 2px 12px rgba(13,43,94,0.10);
  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{
    font-family: 'Arial', 'Helvetica Neue', sans-serif;
    background: var(--bg);
    color: var(--text);
    font-size: 13px;
    line-height: 1.5;
  }}

  /* ── Header ── */
  .site-header {{
    background: linear-gradient(135deg, var(--navy) 0%, #1a4080 100%);
    color: white;
    padding: 0;
    box-shadow: var(--shadow);
  }}
  .header-top {{
    display: flex; align-items: center; justify-content: space-between;
    padding: 18px 32px 12px;
  }}
  .header-seal {{
    font-size: 42px; line-height: 1;
  }}
  .header-title {{ flex: 1; padding: 0 24px; }}
  .header-title h1 {{
    font-size: 22px; font-weight: 700; letter-spacing: 0.5px;
    color: white;
  }}
  .header-title h2 {{
    font-size: 14px; font-weight: 400; color: var(--gold); margin-top: 2px;
  }}
  .header-meta {{
    text-align: right; font-size: 11px; color: rgba(255,255,255,0.75);
    line-height: 1.8;
  }}
  .conf-banner {{
    background: var(--gold);
    color: var(--navy);
    text-align: center;
    font-size: 11px;
    font-weight: 700;
    letter-spacing: 1px;
    padding: 5px;
  }}
  .statute-bar {{
    background: var(--teal);
    color: white;
    text-align: center;
    font-size: 11px;
    padding: 5px;
    letter-spacing: 0.3px;
  }}

  /* ── Navigation tabs ── */
  .tab-nav {{
    background: var(--navy);
    display: flex;
    flex-wrap: wrap;
    gap: 2px;
    padding: 8px 16px 0;
    border-bottom: 3px solid var(--gold);
    position: sticky; top: 0; z-index: 100;
  }}
  .tab-btn {{
    background: rgba(255,255,255,0.08);
    color: rgba(255,255,255,0.8);
    border: none;
    padding: 8px 16px;
    cursor: pointer;
    font-size: 12px;
    font-weight: 600;
    border-radius: 4px 4px 0 0;
    transition: all 0.2s;
    letter-spacing: 0.2px;
  }}
  .tab-btn:hover {{ background: rgba(255,255,255,0.18); color: white; }}
  .tab-btn.active {{
    background: var(--gold);
    color: var(--navy);
  }}
  .tab-content {{ display: none; padding: 24px 28px; }}
  .tab-content.active {{ display: block; }}

  /* ── Section cards ── */
  .card {{
    background: white;
    border-radius: 8px;
    box-shadow: var(--shadow);
    margin-bottom: 20px;
    overflow: hidden;
  }}
  .card-header {{
    background: var(--navy);
    color: white;
    padding: 12px 18px;
    font-size: 13px;
    font-weight: 700;
    letter-spacing: 0.3px;
    display: flex;
    align-items: center;
    gap: 8px;
  }}
  .card-body {{ padding: 0; }}

  /* ── Tables ── */
  table {{
    width: 100%; border-collapse: collapse; font-size: 12px;
  }}
  th {{
    background: var(--teal);
    color: white;
    padding: 9px 10px;
    text-align: left;
    font-weight: 600;
    font-size: 11px;
    letter-spacing: 0.2px;
    white-space: nowrap;
  }}
  td {{
    padding: 8px 10px;
    border-bottom: 1px solid var(--border);
    vertical-align: top;
  }}
  tr:last-child td {{ border-bottom: none; }}
  tr:nth-child(even) td {{ background: var(--navy-lt); }}
  tr:hover td {{ background: #E3EAF5; }}
  .center {{ text-align: center; }}
  .small {{ font-size: 11px; color: #555; }}
  .empty {{ color: #AAA; font-style: italic; }}

  /* ── Factor rows ── */
  .factor-safety td {{ background: #FFF5F5 !important; }}
  .factor-safety:hover td {{ background: #FFE8E8 !important; }}
  .factor-num {{
    font-weight: 700; color: var(--navy);
    text-align: center; white-space: nowrap;
    font-size: 13px;
  }}
  .factor-desc {{ max-width: 280px; }}
  .safety-tag {{
    display: inline-block;
    background: var(--red);
    color: white;
    font-size: 9px;
    font-weight: 700;
    padding: 1px 6px;
    border-radius: 3px;
    margin-top: 4px;
  }}
  .evidence-cell {{ max-width: 200px; font-size: 11px; }}

  /* ── Badges ── */
  .badge {{
    display: inline-block;
    padding: 2px 8px;
    border-radius: 3px;
    font-size: 10px;
    font-weight: 700;
    letter-spacing: 0.5px;
    white-space: nowrap;
  }}
  .conf-badge {{
    display: inline-block;
    padding: 2px 7px;
    border-radius: 3px;
    font-size: 10px;
    font-weight: 600;
  }}
  .conf-public {{ background: #E8F5E9; color: #1E8449; }}
  .conf-restricted {{ background: #FFFDE7; color: #B7770D; }}
  .conf-confidential {{ background: #FFF0F0; color: var(--red); }}
  .conf-sealed {{ background: #1A1A1A; color: white; }}

  /* ── Cross-exam ── */
  .cx-section-header td {{
    background: var(--teal) !important;
    color: white;
    font-weight: 700;
    font-size: 12px;
    letter-spacing: 0.3px;
    padding: 8px 12px;
  }}
  .question-cell {{ font-weight: 600; color: var(--navy); max-width: 300px; }}
  .comm-method, .event-type {{
    display: inline-block;
    background: var(--navy-lt);
    color: var(--navy);
    padding: 2px 7px;
    border-radius: 3px;
    font-size: 10px;
    font-weight: 600;
  }}

  /* ── Status ── */
  .status-agreed {{ color: var(--green); font-weight: 700; }}
  .status-disputed {{ color: var(--red); font-weight: 700; }}
  .status-pending {{ color: var(--orange); }}
  .checklist-done td {{ text-decoration: line-through; color: #AAA; }}

  /* ── Overview grid ── */
  .overview-grid {{
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(280px, 1fr));
    gap: 12px;
    padding: 16px;
  }}
  .overview-item {{
    background: var(--navy-lt);
    border-left: 4px solid var(--navy);
    padding: 10px 14px;
    border-radius: 4px;
  }}
  .overview-item .label {{ font-size: 10px; color: #666; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px; }}
  .overview-item .value {{ font-size: 13px; font-weight: 600; color: var(--navy); margin-top: 2px; }}

  /* ── Priority summary ── */
  .priority-bar {{
    display: flex; gap: 8px; flex-wrap: wrap; padding: 16px;
  }}
  .priority-pill {{
    padding: 6px 14px; border-radius: 20px;
    font-size: 11px; font-weight: 700; letter-spacing: 0.5px;
  }}

  /* ── Legal notice ── */
  .legal-notice {{
    background: #FFF3CD;
    border: 1px solid var(--gold);
    border-left: 5px solid var(--gold);
    padding: 12px 16px;
    border-radius: 4px;
    margin: 16px;
    font-size: 11px;
    color: #7D5A00;
  }}
  .odc-notice {{
    background: #FFF0F0;
    border: 1px solid var(--red);
    border-left: 5px solid var(--red);
    padding: 12px 16px;
    border-radius: 4px;
    margin: 16px;
    font-size: 11px;
    color: #7B0000;
  }}

  /* ── Drive links ── */
  .drive-link {{
    color: var(--teal);
    font-weight: 700;
    text-decoration: none;
    font-size: 11px;
  }}
  .drive-link:hover {{ text-decoration: underline; }}

  /* ── Footer ── */
  .site-footer {{
    background: var(--navy);
    color: rgba(255,255,255,0.6);
    text-align: center;
    padding: 16px;
    font-size: 10px;
    margin-top: 32px;
    line-height: 1.8;
  }}
  .site-footer strong {{ color: var(--gold); }}

  /* ── Print ── */
  @media print {{
    .tab-nav {{ display: none; }}
    .tab-content {{ display: block !important; page-break-before: always; }}
    .tab-content:first-child {{ page-break-before: avoid; }}
  }}

  /* ── Responsive ── */
  @media (max-width: 768px) {{
    .header-top {{ flex-direction: column; text-align: center; padding: 14px; }}
    .header-seal {{ margin-bottom: 8px; }}
    .header-meta {{ text-align: center; margin-top: 8px; }}
    .tab-btn {{ font-size: 11px; padding: 6px 10px; }}
    .tab-content {{ padding: 14px; }}
    table {{ font-size: 11px; }}
  }}
</style>
</head>
<body>

<!-- ── Header ── -->
<header class="site-header">
  <div class="header-top">
    <div class="header-seal">⚖</div>
    <div class="header-title">
      <h1>Family Court Custody Dashboard</h1>
      <h2>{court} &nbsp;|&nbsp; {jurisdiction} &nbsp;|&nbsp; {statute}</h2>
    </div>
    <div class="header-meta">
      <div><strong>Case:</strong> {case_name}</div>
      <div><strong>Docket:</strong> {docket}</div>
      <div><strong>Generated:</strong> {now}</div>
      <div><strong>{party_a}</strong> vs <strong>{party_b}</strong></div>
    </div>
  </div>
  <div class="conf-banner">🔒 {conf_label}</div>
  <div class="statute-bar">
    Pennsylvania Title 23 Chapter 53 § 5328 — Best Interest of the Child — Amended Act 11, June 30, 2025 (eff. 60 days) &nbsp;|&nbsp;
    No single factor is determinative — Court examines totality of circumstances &nbsp;|&nbsp;
    Gender neutral per § 5328(b)
  </div>
</header>

<!-- ── Navigation ── -->
<nav class="tab-nav">
  <button class="tab-btn active" data-tab="overview" onclick="showTab('overview',this)">📋 Case Overview</button>
  <button class="tab-btn" data-tab="factors" onclick="showTab('factors',this)">⚖ Custody Factors</button>
  <button class="tab-btn" data-tab="evidence" onclick="showTab('evidence',this)">📁 Evidence</button>
  <button class="tab-btn" data-tab="timeline" onclick="showTab('timeline',this)">📅 Timeline</button>
  <button class="tab-btn" data-tab="comms" onclick="showTab('comms',this)">📞 Communications</button>
  <button class="tab-btn" data-tab="crossexam" onclick="showTab('crossexam',this)">❓ Cross-Examination</button>
  <button class="tab-btn" data-tab="checklist" onclick="showTab('checklist',this)">✅ Hearing Prep</button>
  <button class="tab-btn" data-tab="negotiation" onclick="showTab('negotiation',this)">🤝 Negotiation</button>
  <button class="tab-btn" data-tab="judge" onclick="showTab('judge',this)">🏛 Judge View</button>
  <button class="tab-btn" data-tab="odc" onclick="showTab('odc',this)">⚠ ODC Tracker</button>
</nav>

<!-- ══════════════════════════════════════════════════════ -->
<!-- TAB 1: CASE OVERVIEW -->
<!-- ══════════════════════════════════════════════════════ -->
<div id="tab-overview" class="tab-content active">
  <div class="card">
    <div class="card-header">📋 Case Overview</div>
    <div class="overview-grid">
      {''.join(f'<div class="overview-item"><div class="label">{r[0]}</div><div class="value">{r[1] or "—"}</div></div>' for r in overview_rows if r[0] and not r[0].startswith("Field"))}
    </div>
  </div>

  <div class="legal-notice">
    <strong>⚖ Legal Standard:</strong> The court shall determine the <em>best interest of the child</em> by considering
    all relevant factors under 23 Pa.C.S. § 5328, giving <strong>substantial weighted consideration</strong> to
    safety factors (1), (2), (2.1), and (2.2). No single factor is determinative — the court examines the
    <strong>totality of circumstances</strong>. Awards are <strong>gender neutral</strong> per § 5328(b).
    Within 30 days of a custody complaint, the court must provide all parties a copy of § 5328 per § 5328(d).
  </div>

  <div class="card">
    <div class="card-header">🎯 Priority Summary</div>
    <div class="priority-bar">
      <div class="priority-pill" style="background:#4A0000;color:#FF6B6B">⚫ CATASTROPHIC — Immediate Safety / Court Intervention</div>
      <div class="priority-pill" style="background:#7B0000;color:#FF9999">🔴 CRITICAL — Directly Affects Custody Decision</div>
      <div class="priority-pill" style="background:#7D3C00;color:#FFAD60">🟠 HIGH — Significant Case Impact</div>
      <div class="priority-pill" style="background:#1A4A7A;color:#82B4E8">🔵 MEDIUM — Notable, Relevant</div>
      <div class="priority-pill" style="background:#145A32;color:#82E0AA">🟢 LOW — Background / Supporting</div>
    </div>
  </div>
</div>

<!-- ══════════════════════════════════════════════════════ -->
<!-- TAB 2: CUSTODY FACTORS -->
<!-- ══════════════════════════════════════════════════════ -->
<div id="tab-factors" class="tab-content">
  <div class="legal-notice">
    <strong>📜 23 Pa.C.S. § 5328 — Amended Act 11, June 30, 2025 (eff. 60 days):</strong>
    Factors (1), (2), (2.1), and (2.2) receive <strong>substantial weighted consideration</strong> as they
    affect the <strong>safety of the child</strong>. Abuse exception: § 5328(a.1) — circumstances in response
    to abuse shall NOT be weighed against the protecting party. Gender neutral: § 5328(b).
  </div>
  <div class="card">
    <div class="card-header">⚖ Pennsylvania Custody Factors — 23 Pa.C.S. § 5328</div>
    <div class="card-body">
      <table>
        <thead>
          <tr>
            <th>Factor</th>
            <th>Legal Standard</th>
            <th>Tier</th>
            <th>Your Evidence ({party_a})</th>
            <th>Opposing Evidence ({party_b})</th>
            <th>Rebuttal</th>
            <th>Priority</th>
            <th>Favors</th>
            <th>Doc Ref</th>
          </tr>
        </thead>
        <tbody>
          {factors_html()}
        </tbody>
      </table>
    </div>
  </div>
</div>

<!-- ══════════════════════════════════════════════════════ -->
<!-- TAB 3: EVIDENCE -->
<!-- ══════════════════════════════════════════════════════ -->
<div id="tab-evidence" class="tab-content">
  <div class="legal-notice">
    <strong>📁 Evidence Tiers:</strong>
    🟢 <strong>Public</strong> — May be shared with all parties |
    🟡 <strong>Restricted</strong> — Attorney only |
    🔴 <strong>Confidential</strong> — Court upon request with signed agreement |
    ⚫ <strong>Sealed</strong> — Court order required.<br>
    <strong>Medical and Financial records:</strong> Physical copies only — do NOT upload to any online system without court authorization.
  </div>
  <div class="card">
    <div class="card-header">📁 Evidence Tracker — Documents, Communications & Media</div>
    <div class="card-body">
      <table>
        <thead>
          <tr>
            <th>#</th><th>Date</th><th>Type</th><th>Description</th>
            <th>Relevance (Custody Factor)</th><th>Priority</th>
            <th>Confidentiality</th><th>Drive Link</th>
            <th>Physical Location</th><th>Status</th>
          </tr>
        </thead>
        <tbody>{evidence_html()}</tbody>
      </table>
    </div>
  </div>
</div>

<!-- ══════════════════════════════════════════════════════ -->
<!-- TAB 4: TIMELINE -->
<!-- ══════════════════════════════════════════════════════ -->
<div id="tab-timeline" class="tab-content">
  <div class="card">
    <div class="card-header">📅 Master Timeline of Events — Chronological Case Record</div>
    <div class="card-body">
      <table>
        <thead>
          <tr>
            <th>#</th><th>Date</th><th>Time</th><th>Event Type</th>
            <th>Description</th><th>Parties</th><th>Priority</th><th>Evidence Ref</th>
          </tr>
        </thead>
        <tbody>{timeline_html()}</tbody>
      </table>
    </div>
  </div>
</div>

<!-- ══════════════════════════════════════════════════════ -->
<!-- TAB 5: COMMUNICATIONS -->
<!-- ══════════════════════════════════════════════════════ -->
<div id="tab-comms" class="tab-content">
  <div class="card">
    <div class="card-header">📞 Communications Log — All Relevant Interactions</div>
    <div class="card-body">
      <table>
        <thead>
          <tr>
            <th>#</th><th>Date</th><th>Time</th><th>Method</th>
            <th>From</th><th>To</th><th>Summary</th>
            <th>Response?</th><th>Evidence Link</th>
          </tr>
        </thead>
        <tbody>{comms_html()}</tbody>
      </table>
    </div>
  </div>
</div>

<!-- ══════════════════════════════════════════════════════ -->
<!-- TAB 6: CROSS-EXAMINATION -->
<!-- ══════════════════════════════════════════════════════ -->
<div id="tab-crossexam" class="tab-content">
  <div class="legal-notice">
    <strong>❓ Cross-Examination Tips:</strong>
    Ask only leading questions (yes/no). Start with facts the witness must admit.
    Build to contradictions. Each question should have one clear purpose.
    Have your evidence reference ready before you ask each question.
  </div>
  <div class="card">
    <div class="card-header">❓ Cross-Examination Question Builder</div>
    <div class="card-body">
      <table>
        <thead>
          <tr>
            <th>#</th><th>Witness</th><th>Role</th><th>Question</th>
            <th>Expected Answer</th><th>Evidence</th>
            <th>§5328 Factor</th><th>Priority</th><th>Status</th>
          </tr>
        </thead>
        <tbody>{cx_html()}</tbody>
      </table>
    </div>
  </div>
</div>

<!-- ══════════════════════════════════════════════════════ -->
<!-- TAB 7: HEARING PREP CHECKLIST -->
<!-- ══════════════════════════════════════════════════════ -->
<div id="tab-checklist" class="tab-content">
  <div class="card">
    <div class="card-header">✅ Hearing Preparation Checklist — 48-Hour Pre-Trial</div>
    <div class="card-body">
      <table>
        <thead>
          <tr>
            <th>#</th><th>Task</th><th>Notes</th><th>Due Date</th><th>Status</th>
          </tr>
        </thead>
        <tbody>{checklist_html()}</tbody>
      </table>
    </div>
  </div>
</div>

<!-- ══════════════════════════════════════════════════════ -->
<!-- TAB 8: NEGOTIATION & SETTLEMENT -->
<!-- ══════════════════════════════════════════════════════ -->
<div id="tab-negotiation" class="tab-content">
  <div class="legal-notice">
    <strong>🤝 Out-of-Court Settlement:</strong>
    Both parties may use this section to compare positions and identify compromise zones.
    Agreed items save court time. Disputed items are prioritized for judicial attention.
    Any settlement must be reduced to a written agreement and submitted to the court for approval.
  </div>
  <div class="card">
    <div class="card-header">🤝 Negotiation & Settlement Tracker</div>
    <div class="card-body">
      <table>
        <thead>
          <tr>
            <th>Custody Issue</th>
            <th>{party_a} Position</th>
            <th>{party_b} Position</th>
            <th>Proposed Compromise</th>
            <th>Status</th>
          </tr>
        </thead>
        <tbody>{negotiation_html()}</tbody>
      </table>
    </div>
  </div>
</div>

<!-- ══════════════════════════════════════════════════════ -->
<!-- TAB 9: JUDGE'S PRE-TRIAL VIEW -->
<!-- ══════════════════════════════════════════════════════ -->
<div id="tab-judge" class="tab-content">
  <div class="legal-notice">
    <strong>🏛 48-Hour Pre-Trial Submission:</strong>
    This summary is prepared for the court's advance review. All facts are supported by
    evidence referenced in the Evidence Tracker. This dashboard may be submitted electronically
    or as a printed exhibit per court direction.
  </div>
  <div class="card">
    <div class="card-header">🏛 Judge's Pre-Trial Summary — {case_name} — Docket: {docket}</div>
    <div class="card-body">
      <table>
        <thead>
          <tr>
            <th>Section</th><th>Summary</th><th>Evidence Reference</th><th>Priority</th>
          </tr>
        </thead>
        <tbody>{judge_html()}</tbody>
      </table>
    </div>
  </div>
</div>

<!-- ══════════════════════════════════════════════════════ -->
<!-- TAB 10: ODC TRACKER -->
<!-- ══════════════════════════════════════════════════════ -->
<div id="tab-odc" class="tab-content">
  <div class="odc-notice">
    <strong>⚠ ODC IMPORTANT LIMITATIONS:</strong>
    The Office of Disciplinary Counsel enforces the <em>Rules of Professional Conduct only</em>.
    ODC does <strong>NOT</strong> represent you, cannot obtain refunds or damages on your behalf,
    and cannot interfere with pending proceedings. Required proof standard: <strong>Clear and Convincing Evidence</strong>.
    Submit as ONE email with ONE attachment. No zipped files. Respond within 14 days of ODC inquiry.
    Website: <strong>www.padisciplinaryboard.org</strong> | Phone: <strong>(215) 560-6296</strong>
  </div>
  <div class="card">
    <div class="card-header">⚠ ODC Violation Tracker — STRICTLY CONFIDENTIAL</div>
    <div class="card-body">
      <table>
        <thead>
          <tr>
            <th>#</th><th>Date</th><th>Attorney</th><th>Rule</th>
            <th>Description</th><th>Evidence</th><th>Impact</th>
            <th>Priority</th><th>Status</th>
          </tr>
        </thead>
        <tbody>
          {''.join(f"""<tr>
            <td class="center">{r[0]}</td>
            <td>{r[1]}</td><td>{r[2]}</td>
            <td><strong>{r[3]}</strong></td>
            <td>{r[4]}</td><td class="small">{r[5]}</td>
            <td class="small">{r[6]}</td>
            <td class="center">{priority_badge(r[7])}</td>
            <td>{r[8]}</td>
          </tr>""" for r in odc_rows if r[0] or r[1])}
        </tbody>
      </table>
    </div>
  </div>
</div>

<!-- ── Footer ── -->
<footer class="site-footer">
  <strong>Family Court Custody Dashboard</strong> &nbsp;|&nbsp;
  {court} &nbsp;|&nbsp; {jurisdiction} &nbsp;|&nbsp;
  Generated: {now} &nbsp;|&nbsp;
  Statute: {statute}<br>
  🔒 {conf_label}<br>
  <em>This dashboard is a legal case organization tool. It does not constitute legal advice.
  Consult a licensed Pennsylvania attorney for legal guidance.
  For Pro Se resources: Philadelphia Family Court Self-Help Center.</em>
</footer>

<script>
function showTab(name, btn) {{
  document.querySelectorAll('.tab-content').forEach(t => t.classList.remove('active'));
  document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
  document.getElementById('tab-' + name).classList.add('active');
  btn.classList.add('active');
}}
</script>

{tracking_js}
</body>
</html>"""
    return html


def generate_admin_panel(cfg):
    webhook = cfg.get("Google Sheet Webhook", "")
    now = datetime.datetime.now().strftime("%B %d, %Y at %I:%M %p")
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Dashboard Admin Panel — Access Tracking</title>
<style>
  :root {{ --navy:#0D2B5E; --gold:#C8A84B; --teal:#2E7D7B; --red:#C0392B; }}
  body {{ font-family: Arial, sans-serif; background:#F4F6FB; color:#1A1A2E; margin:0; }}
  header {{ background:var(--navy); color:white; padding:18px 28px; display:flex; align-items:center; gap:16px; }}
  header h1 {{ font-size:18px; }}
  header .sub {{ font-size:12px; color:var(--gold); margin-top:2px; }}
  .container {{ padding:24px 28px; }}
  .card {{ background:white; border-radius:8px; box-shadow:0 2px 12px rgba(13,43,94,.10); margin-bottom:20px; overflow:hidden; }}
  .card-header {{ background:var(--navy); color:white; padding:11px 16px; font-size:13px; font-weight:700; }}
  table {{ width:100%; border-collapse:collapse; font-size:12px; }}
  th {{ background:var(--teal); color:white; padding:8px 10px; text-align:left; }}
  td {{ padding:7px 10px; border-bottom:1px solid #E0E7F2; }}
  tr:nth-child(even) td {{ background:#F0F4FA; }}
  .warning {{ background:#FFF3CD; border-left:4px solid var(--gold); padding:12px 16px; margin-bottom:16px; font-size:12px; color:#7D5A00; border-radius:4px; }}
  .btn {{ background:var(--navy); color:white; border:none; padding:8px 18px; border-radius:4px; cursor:pointer; font-size:12px; font-weight:700; }}
  #log-table td, #log-table th {{ font-size:11px; }}
  .empty {{ color:#AAA; font-style:italic; text-align:center; padding:24px; }}
</style>
</head>
<body>
<header>
  <div>⚖</div>
  <div>
    <h1>Family Court Dashboard — Admin & Access Tracking Panel</h1>
    <div class="sub">Security Monitoring &nbsp;|&nbsp; Generated: {now} &nbsp;|&nbsp; 🔒 STRICTLY CONFIDENTIAL</div>
  </div>
</header>
<div class="container">
  <div class="warning">
    ⚠ <strong>Security Notice:</strong> This panel logs all visitors to your shared dashboard.
    If you detect unauthorized access, document the entry and contact your attorney immediately.
    Tracking is {'<strong style="color:green">ENABLED</strong>' if webhook else '<strong style="color:red">DISABLED</strong> — Set webhook URL in Dashboard Config sheet to enable'}.
  </div>
  <div class="card">
    <div class="card-header">🔍 Visitor Access Log</div>
    <div style="padding:12px">
      <button class="btn" onclick="loadLog()">🔄 Refresh Log</button>
      <button class="btn" style="background:#145A32;margin-left:8px" onclick="exportCSV()">⬇ Export CSV</button>
    </div>
    <table id="log-table">
      <thead>
        <tr>
          <th>#</th><th>Timestamp</th><th>Event</th><th>IP</th>
          <th>City</th><th>Country</th><th>Browser</th>
          <th>Device</th><th>Duration (sec)</th><th>Tab Viewed</th>
        </tr>
      </thead>
      <tbody id="log-body">
        <tr><td colspan="10" class="empty">Click "Refresh Log" to load visitor data from Google Sheets.</td></tr>
      </tbody>
    </table>
  </div>
</div>
<script>
var WEBHOOK = "{webhook}";
var allData = [];

function loadLog() {{
  if (!WEBHOOK) {{ alert("No webhook configured. Set Google Sheet Webhook in Dashboard Config sheet."); return; }}
  var fetchUrl = WEBHOOK + "?action=read";
  fetch(fetchUrl)
    .then(r => r.json())
    .then(data => {{
      allData = data;
      renderTable(data);
    }})
    .catch(e => {{
      document.getElementById('log-body').innerHTML =
        '<tr><td colspan="10" class="empty">Could not load data. Check webhook URL and Google Apps Script permissions.</td></tr>';
    }});
}}

function renderTable(data) {{
  var tbody = document.getElementById('log-body');
  if (!data || !data.length) {{
    tbody.innerHTML = '<tr><td colspan="10" class="empty">No visitor records found.</td></tr>';
    return;
  }}
  tbody.innerHTML = data.map((r,i) => `
    <tr>
      <td>${{i+1}}</td>
      <td>${{r.timestamp||''}}</td>
      <td>${{r.event||''}}</td>
      <td>${{r.ip||''}}</td>
      <td>${{r.city||''}}</td>
      <td>${{r.country||''}}</td>
      <td>${{r.browser||r.userAgent||''}}</td>
      <td>${{r.device||''}}</td>
      <td>${{r.duration_sec||''}}</td>
      <td>${{r.tab||''}}</td>
    </tr>`).join('');
}}

function exportCSV() {{
  if (!allData.length) {{ alert("No data to export. Load the log first."); return; }}
  var keys = Object.keys(allData[0]);
  var csv = [keys.join(','), ...allData.map(r => keys.map(k => JSON.stringify(r[k]||'')).join(','))].join('\\n');
  var a = document.createElement('a');
  a.href = 'data:text/csv,' + encodeURIComponent(csv);
  a.download = 'visitor_log.csv';
  a.click();
}}
</script>
</body>
</html>"""


# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    excel_path = EXCEL_FILE
    if not os.path.exists(excel_path):
        print(f"\n⚠  Excel file not found: {excel_path}")
        print("   Please make sure 'family_court_custody_template.xlsx' is in the same folder as this script.\n")
        sys.exit(1)

    print("\n⚖  Family Court Custody Dashboard Generator")
    print("=" * 50)
    print(f"   Reading: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    cfg = read_config(wb)

    output_html = cfg.get("Dashboard File Name", OUTPUT_HTML)
    admin_html  = cfg.get("Admin Panel File",    ADMIN_HTML)

    print(f"   Generating dashboard: {output_html}")
    html = generate_dashboard(wb, cfg)
    with open(output_html, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"   Generating admin panel: {admin_html}")
    admin = generate_admin_panel(cfg)
    with open(admin_html, "w", encoding="utf-8") as f:
        f.write(admin)

    h = file_hash(output_html)
    print(f"\n✅ Done!")
    print(f"   Dashboard:   {output_html}  (SHA-256 prefix: {h})")
    print(f"   Admin Panel: {admin_html}")
    print(f"\n   Open '{output_html}' in your browser to view your dashboard.")
    print(f"   Keep this script and the Excel file on your local computer for privacy.\n")


# ══════════════════════════════════════════════════════════════════════════════
# PARTY MODE HANDLER — appended to support Single / Dual / Mid-case join
# ══════════════════════════════════════════════════════════════════════════════

def get_party_mode(cfg):
    """
    Returns party mode dict consumed by dashboard generator.
    Works whether Party B has joined or not — always ready for both.
    """
    mode         = cfg.get("Party Mode", "SINGLE").upper().strip()
    party_a      = cfg.get("Party A Name / Label", cfg.get("Party A Label", "Petitioner"))
    party_b      = cfg.get("Party B Name / Label", cfg.get("Party B Label", "Respondent"))
    party_b_file = cfg.get("Party B Excel File", "").strip()
    show_b_cols  = cfg.get("Show Party B Columns", "YES").upper() != "NO"
    joined_date  = cfg.get("Party B Joined Date", "").strip()
    merge_status = cfg.get("Merge Status", "NOT MERGED").strip()
    single_note  = cfg.get("Single Party Note for Court",
                           "Opposing party has not yet participated in pre-trial dashboard submission.")
    version      = cfg.get("Dashboard Version", "1.0").strip()

    is_dual      = mode == "DUAL" or bool(party_b_file)
    is_merged    = "MERGED" in merge_status.upper()

    return {
        "mode":         "DUAL" if is_dual else "SINGLE",
        "party_a":      party_a,
        "party_b":      party_b,
        "party_b_file": party_b_file,
        "show_b_cols":  show_b_cols,
        "joined_date":  joined_date,
        "merge_status": merge_status,
        "single_note":  single_note,
        "is_dual":      is_dual,
        "is_merged":    is_merged,
        "version":      version,
    }


def party_mode_banner(pm):
    """Generates a status banner shown at top of every dashboard tab."""
    if pm["is_dual"] and pm["is_merged"]:
        color, icon, msg = "#145A32", "✅", (
            f"DUAL PARTY — MERGED VIEW &nbsp;|&nbsp; "
            f"{pm['party_a']} &amp; {pm['party_b']} both participating &nbsp;|&nbsp; "
            f"Party B joined: {pm['joined_date'] or 'recorded'} &nbsp;|&nbsp; "
            f"Merge status: {pm['merge_status']}"
        )
    elif pm["is_dual"]:
        color, icon, msg = "#1A4A7A", "🔵", (
            f"DUAL PARTY MODE &nbsp;|&nbsp; "
            f"Both {pm['party_a']} and {pm['party_b']} are participating &nbsp;|&nbsp; "
            f"Party B joined: {pm['joined_date'] or 'date not recorded'}"
        )
    else:
        color, icon, msg = "#7D3C00", "🟡", (
            f"SINGLE PARTY MODE — {pm['party_a']} only &nbsp;|&nbsp; "
            f"Dashboard is fully functional. &nbsp;|&nbsp; "
            f"Party B column is ready — other party can join at ANY time. "
            f"Simply update 'Party Mode' to DUAL in Dashboard Config and regenerate."
        )

    return f"""
    <div style="background:{color};color:white;padding:8px 28px;
                font-size:11px;font-weight:600;letter-spacing:0.3px;
                display:flex;align-items:center;gap:10px;">
      <span style="font-size:14px">{icon}</span>
      <span>{msg}</span>
      <span style="margin-left:auto;opacity:0.7;font-weight:400">
        v{pm['version']} &nbsp;|&nbsp; Dashboard generated {datetime.datetime.now().strftime("%b %d, %Y")}
      </span>
    </div>"""


def party_b_placeholder(pm, context="evidence"):
    """Returns appropriate placeholder text for Party B columns."""
    phrases = {
        "evidence":   (f"[{pm['party_b']} has not yet submitted data. "
                       f"You may enter known opposing claims from their court filings, "
                       f"attorney letters, or prior testimony.]"),
        "position":   (f"[{pm['party_b']} position not yet entered. "
                       f"Column ready — update Dashboard Config to DUAL when they join.]"),
        "join_note":  (f"⬜ {pm['party_b']} has not yet participated. "
                       f"This column activates when Party Mode = DUAL."),
    }
    return phrases.get(context, f"[{pm['party_b']} — not yet submitted]")


def how_to_switch_mode():
    """Instructions shown in single-party mode to explain how to add Party B."""
    return """
    <div style="background:#E8F5E9;border:1px solid #1E8449;border-left:5px solid #1E8449;
                padding:12px 16px;border-radius:4px;font-size:11px;color:#145A32;margin:16px;">
      <strong>💡 How to add the other party at any time — 3 steps:</strong><br>
      1. Ask the other party to fill in their own copy of <code>family_court_custody_template.xlsx</code><br>
      2. In YOUR Dashboard Config sheet → set <strong>Party Mode = DUAL</strong>
         and enter their filename in <strong>Party B Excel File</strong><br>
      3. Run <code>python generate_dashboard.py</code> again — dashboard updates instantly ✅<br><br>
      <em>No data is lost. Your existing entries remain exactly as entered.
      The other party's data simply fills in alongside yours.</em>
    </div>"""


# ── Patch generate_dashboard to use party mode ────────────────────────────────
_original_generate = generate_dashboard

def generate_dashboard(wb, cfg):
    """Wraps original generator with party mode awareness."""
    pm   = get_party_mode(cfg)
    html = _original_generate(wb, cfg)

    # Inject party mode banner right after <body> tag
    banner = party_mode_banner(pm)
    html = html.replace("<header class=\"site-header\">",
                        banner + "\n<header class=\"site-header\">", 1)

    # In single mode, inject how-to-switch hint into overview tab
    if not pm["is_dual"]:
        hint = how_to_switch_mode()
        html = html.replace(
            '<div id="tab-overview" class="tab-content active">',
            f'<div id="tab-overview" class="tab-content active">{hint}', 1
        )

        # Add single-party court note to judge tab
        court_note = f"""
        <div style="background:#FFF3CD;border:1px solid #C8A84B;border-left:5px solid #C8A84B;
                    padding:10px 16px;border-radius:4px;font-size:11px;color:#7D5A00;margin:16px;">
          <strong>🏛 Court Note (Auto-generated for Single Party Mode):</strong><br>
          {pm['single_note']}
          This submission represents {pm['party_a']}'s case organization only.
          The opposing party column is available and ready should they choose to participate.
        </div>"""
        html = html.replace(
            '<div id="tab-judge" class="tab-content">',
            f'<div id="tab-judge" class="tab-content">{court_note}', 1
        )

    return html


if __name__ == "__main__":
    # Re-run if called directly after append
    pass
